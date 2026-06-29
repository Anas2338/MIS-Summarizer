"""Core conversion logic for MIS Data Converter.

Reads a raw MIS Excel export, cleans it, groups by financial year and
section, and writes a formatted output matching the exact layout of
the reference "final product.xls".
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def read_input(input_path: str) -> pd.DataFrame:
    """Read the raw Excel file and return a DataFrame with stripped column names."""
    df = pd.read_excel(input_path)
    df.columns = df.columns.str.strip()
    return df


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Clean the DataFrame: drop columns, parse types.

    Keeps only: Wa Name, Section, Tax Year, Taxable Amount, Paid Amount, Payment Date
    Converts Paid Amount from comma-string to int.
    Parses Payment Date to datetime.
    """
    df = df.copy()
    df.columns = df.columns.str.strip()

    # Keep only the 6 columns used in the final output
    keep_cols = [
        "Wa Name",
        "Section",
        "Tax Year",
        "Taxable Amount",
        "Paid Amount",
        "Payment Date",
    ]
    df = df[[c for c in keep_cols if c in df.columns]]

    # Convert Paid Amount: remove commas, convert to int
    df["Paid Amount"] = (
        df["Paid Amount"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    df["Paid Amount"] = (
        pd.to_numeric(df["Paid Amount"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    # Parse Payment Date (format: 14-Jul-2015)
    df["Payment Date"] = pd.to_datetime(
        df["Payment Date"], format="%d-%b-%Y", errors="coerce"
    )

    # Ensure Taxable Amount is numeric
    df["Taxable Amount"] = (
        pd.to_numeric(df["Taxable Amount"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    # Ensure Tax Year is int
    if "Tax Year" in df.columns:
        df["Tax Year"] = df["Tax Year"].fillna(0).astype(int)

    return df


def add_financial_year(df: pd.DataFrame) -> pd.DataFrame:
    """Add Financial Year label and sort key derived from Payment Date.

    Financial year runs from 1 July to 30 June.
    e.g. 14-Jul-2015 -> TAX YEAR 2015-16, 08-Jan-2016 -> TAX YEAR 2015-16
    """
    df = df.copy()

    def _fy_label(date):
        if pd.isna(date):
            return "TAX YEAR UNKNOWN"
        year = date.year
        month = date.month
        if month >= 7:
            return f"TAX YEAR {year}-{str(year + 1)[-2:]}"
        else:
            return f"TAX YEAR {year - 1}-{str(year)[-2:]}"

    def _fy_sort_key(date):
        """Return the starting year of the FY for sorting."""
        if pd.isna(date):
            return 9999
        year = date.year
        month = date.month
        return year if month >= 7 else year - 1

    df["Financial Year"] = df["Payment Date"].apply(_fy_label)
    df["_fy_sort"] = df["Payment Date"].apply(_fy_sort_key)

    return df


def sort_data(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by financial year -> Section -> Payment Date."""
    df = df.copy()
    df = df.sort_values(
        by=["_fy_sort", "Section", "Payment Date"],
        ascending=[True, True, True],
    ).reset_index(drop=True)
    return df


def write_output(df: pd.DataFrame, output_path: str) -> None:
    """Write a formatted Excel workbook.

    Layout:
      Row 1: Column headers: Wa Name, Section, Tax Year , Taxable Amount,
             Paid Amount, Payment Date
      Then for each financial year:
        - "TAX YEAR XXXX-XX" heading (merged, centred across all columns)
        - Data rows sorted by Section -> Payment Date
        - Sections with >1 row get a subtotal (SUM formula in cols D & E,
          rest blank)
        - 1 blank row between sections within a FY
        - 2 blank rows between FY groups
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Total Paid Amount"

    # ── Styles ──────────────────────────────────────────────────────────
    header_font = Font(bold=True, size=11)
    year_font = Font(bold=True, size=12)
    year_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(size=11)
    subtotal_font = Font(bold=True, size=12)  # one size larger than data
    center_align = Alignment(horizontal="center", vertical="center")
    accounting_format = '#,##0'

    # Border styles: thick (outside) and thin (inside)
    thick_side = Side(style="medium")
    thin_side = Side(style="thin")
    outer_border = Border(
        top=thick_side, bottom=thick_side,
        left=thick_side, right=thick_side,
    )
    header_border = Border(
        top=thick_side, bottom=thick_side,
        left=thin_side, right=thin_side,
    )

    # ── Columns (6 cols) ────────────────────────────────────────────────
    display_columns = [
        "Wa Name",
        "Section",
        "Tax Year ",
        "Taxable Amount",
        "Paid Amount",
        "Payment Date",
    ]
    num_cols = len(display_columns)

    # ── Prepare output DataFrame ────────────────────────────────────────
    out = df.copy()
    out["Payment Date"] = out["Payment Date"].dt.strftime("%d-%b-%Y")
    out["Tax Year"] = out["Tax Year"].astype(int)

    current_row = 1

    # ── Row 1: Column headers ───────────────────────────────────────────
    for col_idx, col_name in enumerate(display_columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        # Dark outside / light inside border
        left = thick_side if col_idx == 1 else thin_side
        right = thick_side if col_idx == num_cols else thin_side
        cell.border = Border(
            top=thick_side, bottom=thick_side,
            left=left, right=right,
        )
    current_row += 1

    # ── Group by financial year ─────────────────────────────────────────
    fy_groups = out.groupby("_fy_sort", sort=False)
    fy_count = len(fy_groups)
    processed = 0

    for fy_sort_key, fy_group in fy_groups:
        processed += 1
        fy_label = fy_group["Financial Year"].iloc[0]

        # --- Financial year heading (merged, dark outline border) ---
        merge_range = (
            f"A{current_row}:{get_column_letter(num_cols)}{current_row}"
        )
        # Set full thick border on all cells first, then merge
        for col_idx in range(1, num_cols + 1):
            ws.cell(row=current_row, column=col_idx).border = outer_border
        ws.merge_cells(merge_range)
        cell = ws.cell(row=current_row, column=1, value=fy_label)
        cell.font = year_font
        cell.alignment = year_alignment
        current_row += 1

        # --- Group by Section within this FY (preserving sorted order) ---
        section_groups = fy_group.groupby("Section", sort=False)
        section_list = list(section_groups)
        num_sections = len(section_list)

        for sec_idx, (section, sec_group) in enumerate(section_list):
            num_rows = len(sec_group)
            first_data_row = current_row  # row where this section's data starts

            # Write data rows
            for _, row in sec_group.iterrows():
                ws.cell(
                    row=current_row, column=1, value=row["Wa Name"]
                )
                ws.cell(
                    row=current_row, column=2, value=row["Section"]
                )
                ws.cell(
                    row=current_row, column=3, value=row["Tax Year"]
                )
                ws.cell(
                    row=current_row, column=4, value=row["Taxable Amount"]
                )
                ws.cell(
                    row=current_row, column=5, value=row["Paid Amount"]
                )
                ws.cell(
                    row=current_row, column=6, value=row["Payment Date"]
                )
                for c in range(1, num_cols + 1):
                    ws.cell(row=current_row, column=c).font = data_font
                    ws.cell(row=current_row, column=c).alignment = center_align
                # Accounting format for amount columns
                ws.cell(row=current_row, column=4).number_format = accounting_format
                ws.cell(row=current_row, column=5).number_format = accounting_format
                # Single-row sections: bold + larger font for amounts
                if num_rows == 1:
                    for c in [4, 5]:
                        cell = ws.cell(row=current_row, column=c)
                        cell.font = subtotal_font
                current_row += 1

            # --- Subtotal (only for sections with >1 row) ---
            # Uses Excel SUM formula so users can verify totals
            if num_rows > 1:
                last_data_row = current_row - 1
                d_formula = f"=SUM(D{first_data_row}:D{last_data_row})"
                e_formula = f"=SUM(E{first_data_row}:E{last_data_row})"

                ws.cell(row=current_row, column=4, value=d_formula)
                ws.cell(row=current_row, column=5, value=e_formula)
                for c in [4, 5]:
                    cell = ws.cell(row=current_row, column=c)
                    cell.font = subtotal_font
                    cell.alignment = center_align
                    cell.number_format = accounting_format
                current_row += 1

            # --- Blank row between sections (only if more sections follow) ---
            if sec_idx < num_sections - 1:
                current_row += 1

        # --- Blank rows between FY groups ---
        if processed < fy_count:
            current_row += 2

    # ── Column widths ───────────────────────────────────────────────────
    # Auto-fit Section (B) and Wa Name (A) based on longest value
    max_section_len = len("Section")  # at least header length
    max_wa_len = len("Wa Name")
    for _, row in out.iterrows():
        max_section_len = max(max_section_len, len(str(row["Section"])))
        max_wa_len = max(max_wa_len, len(str(row["Wa Name"])))

    col_widths = {
        "A": min(max_wa_len + 3, 55),   # Wa Name
        "B": min(max_section_len + 3, 65),  # Section (dynamic)
        "C": 14,   # Tax Year
        "D": 18,   # Taxable Amount
        "E": 16,   # Paid Amount
        "F": 18,   # Payment Date
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)


def convert(input_path: str, output_path: str) -> None:
    """Main orchestrator: read -> clean -> add_fy -> sort -> write."""
    df = read_input(input_path)
    df = clean_data(df)
    df = add_financial_year(df)
    df = sort_data(df)
    write_output(df, output_path)
