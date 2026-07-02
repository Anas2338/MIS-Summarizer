"""Sale/Purchase invoice summarizer.

Reads raw sale/purchase MIS Excel data, removes extraneous columns,
groups invoices by buyer name, inserts subtotal rows, and writes
a formatted Excel output matching the reference layout.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


# ── Columns kept in the output ──────────────────────────────────────────
KEEP_COLS = [
    "Invoice No.",
    "Invoice Type",
    "Invoice Date",
    "Buyer Name",
    "Seller Name",
    "Quantity",
    "Rate",
    "Value of Sales Excluding Sales Tax",
    "Sales Tax/ FED in ST Mode",
    "Further Tax",
]

# Numeric columns that need parsing
NUMERIC_COLS = [
    "Quantity",
    "Rate",
    "Value of Sales Excluding Sales Tax",
    "Sales Tax/ FED in ST Mode",
    "Further Tax",
]

# Columns whose values are summed in the TOTAL row
SUM_COLS = [
    "Value of Sales Excluding Sales Tax",
    "Sales Tax/ FED in ST Mode",
    "Further Tax",
]


# ── Tax year helpers (for "Tax Year" sort mode) ────────────────────────────

def _tax_year_label(date) -> str:
    """Derive a financial-year label from a datetime.

    Financial year runs 1 July – 30 June.
    25-May-2026 → "TAX YEAR 2025-26"
    14-Jul-2025 → "TAX YEAR 2025-26"
    """
    if pd.isna(date):
        return "TAX YEAR UNKNOWN"
    year = date.year
    month = date.month
    if month >= 7:
        return f"TAX YEAR {year}-{str(year + 1)[-2:]}"
    else:
        return f"TAX YEAR {year - 1}-{str(year)[-2:]}"


def _tax_year_sort_key(date) -> int:
    """Return the starting year of the financial year (for sorting)."""
    if pd.isna(date):
        return 9999
    year = date.year
    month = date.month
    return year if month >= 7 else year - 1


def read_input(input_path: str) -> pd.DataFrame:
    """Read the raw Excel file and return a DataFrame with stripped column names."""
    df = pd.read_excel(input_path)
    df.columns = df.columns.str.strip()
    return df


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only the 10 required columns and parse numeric/date types.

    Drops 21 columns that are not needed for the sale/purchase summary.
    Converts Quantity, Rate, and the three amount columns to numeric.
    Parses Invoice Date to datetime.
    """
    df = df.copy()
    df.columns = df.columns.str.strip()

    # Keep only columns that exist in the input
    existing = [c for c in KEEP_COLS if c in df.columns]
    df = df[existing]

    # Parse numeric columns — raw values may have commas (e.g. "22,080.00")
    # or percent signs (e.g. "25%") that need to be stripped first.
    FURTHER_TAX = "Further Tax"
    for col in NUMERIC_COLS:
        if col in df.columns:
            s = df[col].astype(str).str.strip()
            # Strip commas used as thousands separators
            s = s.str.replace(",", "", regex=False)
            # Rate may be stored as "25%" — strip percent and divide by 100
            if col == "Rate":
                s = s.str.replace("%", "", regex=False)
                val = pd.to_numeric(s, errors="coerce").fillna(0)
                df[col] = val / 100.0
            elif col == FURTHER_TAX:
                # Keep NaN for Further Tax — only filled in TOTAL rows
                df[col] = pd.to_numeric(s, errors="coerce")
            else:
                df[col] = pd.to_numeric(s, errors="coerce").fillna(0)

    # Parse Invoice Date (formats like 25-May-2026)
    if "Invoice Date" in df.columns:
        df["Invoice Date"] = pd.to_datetime(
            df["Invoice Date"], format="%d-%b-%Y", errors="coerce"
        )

    return df


def group_and_total(df: pd.DataFrame, sort_by: str = "Buyer Name") -> list[dict]:
    """Group invoices by *sort_by* column and insert TOTAL + blank separator rows.

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned invoice data.
    sort_by : str
        Column to group/sort by — ``"Buyer Name"``, ``"Seller Name"``,
        or ``"Tax Year"``.

    Returns an ordered list of row dicts.  Each dict has a ``_row_type`` key
    (``"data"``, ``"total"``, ``"blank"``, or ``"heading"``) so the writer
    can style accordingly.
    """
    # Work out which of the SUM_COLS are actually present
    present_sum_cols = [c for c in SUM_COLS if c in df.columns]
    present_cols = [c for c in KEEP_COLS if c in df.columns]

    rows: list[dict] = []

    # ── Tax Year sort mode ─────────────────────────────────────────────────
    if sort_by == "Tax Year":
        df = df.copy()

        # If Invoice Date is missing, fall back to a single "UNKNOWN" group
        if "Invoice Date" not in df.columns:
            df["_tax_year_label"] = "TAX YEAR UNKNOWN"
            df["_tax_year_sort"] = 9999
        else:
            # Derive tax year label and sort key from Invoice Date
            df["_tax_year_label"] = df["Invoice Date"].apply(_tax_year_label)
            df["_tax_year_sort"] = df["Invoice Date"].apply(_tax_year_sort_key)

        # Sort by tax year → Invoice Date asc → Invoice No. desc
        sort_cols = ["_tax_year_sort"]
        ascending = [True]
        if "Invoice Date" in df.columns:
            sort_cols.append("Invoice Date")
            ascending.append(True)
        if "Invoice No." in df.columns:
            sort_cols.append("Invoice No.")
            ascending.append(False)
        df = df.sort_values(by=sort_cols, ascending=ascending).reset_index(drop=True)

        # Group by tax year label
        groups = df.groupby("_tax_year_label", sort=False)
        group_names = list(groups.groups.keys())

        for idx, (group_name, group) in enumerate(groups):
            # Emit heading row
            rows.append({"_row_type": "heading", "_heading": group_name})

            # Emit data rows
            for _, row in group.iterrows():
                entry: dict = {"_row_type": "data"}
                for col in present_cols:
                    val = row[col]
                    entry[col] = val
                if "Invoice Date" in entry and pd.notna(entry["Invoice Date"]):
                    entry["Invoice Date"] = entry["Invoice Date"].strftime("%d-%b-%Y")
                rows.append(entry)

            # Emit TOTAL row for this tax year
            total_entry: dict = {"_row_type": "total"}
            for col in present_cols:
                if col == "Quantity":
                    total_entry[col] = "TOTAL"
                elif col in present_sum_cols:
                    total_entry[col] = int(group[col].fillna(0).sum())
                else:
                    total_entry[col] = None
            rows.append(total_entry)

            # Emit blank separator (except after the last group)
            if idx < len(group_names) - 1:
                rows.append({"_row_type": "blank"})

        return rows

    # ── Buyer Name / Seller Name sort mode (original behaviour) ─────────────
    # Sort by chosen column (case-insensitive) → Invoice No. descending
    df["_sort_key"] = df[sort_by].str.lower()
    sort_cols = ["_sort_key"]
    ascending = [True]
    if "Invoice No." in df.columns:
        sort_cols.append("Invoice No.")
        ascending.append(False)
    df = df.sort_values(by=sort_cols, ascending=ascending).reset_index(drop=True)

    if sort_by not in df.columns:
        return rows

    groups = df.groupby(sort_by, sort=False)
    group_names = list(groups.groups.keys())

    for idx, (group_name, group) in enumerate(groups):
        # Emit data rows
        for _, row in group.iterrows():
            entry: dict = {"_row_type": "data"}
            for col in present_cols:
                val = row[col]
                # Keep NaN as NaN for individual data rows (matching reference)
                # but store numeric 0 for easier total computation later
                entry[col] = val
            # Format Invoice Date as string for display
            if "Invoice Date" in entry and pd.notna(entry["Invoice Date"]):
                entry["Invoice Date"] = entry["Invoice Date"].strftime("%d-%b-%Y")
            rows.append(entry)

        # Emit TOTAL row
        total_entry: dict = {"_row_type": "total"}
        for col in present_cols:
            if col == "Quantity":
                total_entry[col] = "TOTAL"
            elif col in present_sum_cols:
                total_entry[col] = int(group[col].fillna(0).sum())
            else:
                total_entry[col] = None
        rows.append(total_entry)

        # Emit blank separator row (except after the last buyer)
        if idx < len(group_names) - 1:
            rows.append({"_row_type": "blank"})

    return rows


def write_output(rows: list[dict], output_path: str) -> None:
    """Write a formatted .xlsx workbook matching the FINAL OUTPUT layout.

    Layout:
      Row 1: "SALE/PURCHASE" title (merged, centred, bold)
      Row 2: Column headers (bold, centred)
      Then for each buyer group:
        - Data rows (normal weight)
        - TOTAL row (bold, sums in the three amount columns)
        - Blank separator row
    """
    # ── Determine columns present (from the first data/total row) ──────
    display_columns = [
        "Invoice No.",
        "Invoice Type",
        "Invoice Date",
        "Buyer Name",
        "Seller Name",
        "Quantity",
        "Rate",
        "Value of Sales Excluding Sales Tax",
        "Sales Tax/ FED in ST Mode",
        "Further Tax",
    ]
    # Filter to columns actually in the data
    first_data = next((r for r in rows if r["_row_type"] in ("data", "total")), None)
    if first_data is not None:
        display_columns = [c for c in display_columns if c in first_data]

    num_cols = len(display_columns)

    # Index of amount columns (for number formatting)
    amount_cols = {
        "Value of Sales Excluding Sales Tax",
        "Sales Tax/ FED in ST Mode",
        "Further Tax",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "SALE PURCHASE"

    # ── Styles ──────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11)
    data_font = Font(size=11)
    total_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    accounting_format = "#,##0"

    thick_side = Side(style="medium")
    thin_side = Side(style="thin")
    outer_border = Border(
        top=thick_side,
        bottom=thick_side,
        left=thick_side,
        right=thick_side,
    )

    current_row = 1

    # ── Row 1: Title ────────────────────────────────────────────────────
    merge_range = f"A{current_row}:{get_column_letter(num_cols)}{current_row}"
    for col_idx in range(1, num_cols + 1):
        ws.cell(row=current_row, column=col_idx).border = outer_border
    ws.merge_cells(merge_range)
    cell = ws.cell(row=current_row, column=1, value="SALE/PURCHASE")
    cell.font = title_font
    cell.alignment = center_align
    current_row += 1

    # ── Row 2: Column headers ───────────────────────────────────────────
    for col_idx, col_name in enumerate(display_columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        left = thick_side if col_idx == 1 else thin_side
        right = thick_side if col_idx == num_cols else thin_side
        cell.border = Border(
            top=thick_side,
            bottom=thick_side,
            left=left,
            right=right,
        )
    current_row += 1

    # ── Data / TOTAL / blank rows ───────────────────────────────────────
    # Build a column-name → column-letter map for SUM formulas
    col_letter_map: dict[str, str] = {}
    for idx, col_name in enumerate(display_columns, 1):
        col_letter_map[col_name] = get_column_letter(idx)

    # Columns in the TOTAL row that get borders ("TOTAL" label + Rate + 3 amounts)
    total_border_cols = {
        "Quantity",
        "Rate",
        "Value of Sales Excluding Sales Tax",
        "Sales Tax/ FED in ST Mode",
        "Further Tax",
    }

    # Track the Excel row range of data rows for the current buyer group
    buyer_data_start: int | None = None
    buyer_data_end: int | None = None

    # Track all TOTAL row numbers (for the GRAND TOTAL formula at the end)
    total_row_nums: list[int] = []

    for row_dict in rows:
        rtype = row_dict["_row_type"]

        if rtype == "blank":
            current_row += 1
            continue

        if rtype == "heading":
            # Single merged cell spanning all columns (e.g. "TAX YEAR 2025-26")
            heading_text = row_dict.get("_heading", "")
            merge_range = (
                f"A{current_row}:{get_column_letter(num_cols)}{current_row}"
            )
            for col_idx in range(1, num_cols + 1):
                ws.cell(row=current_row, column=col_idx).border = outer_border
            ws.merge_cells(merge_range)
            cell = ws.cell(row=current_row, column=1, value=heading_text)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            continue

        if rtype == "data":
            # Track data row range for the current buyer group
            if buyer_data_start is None:
                buyer_data_start = current_row
            buyer_data_end = current_row

            for col_idx, col_name in enumerate(display_columns, 1):
                val = row_dict.get(col_name)
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    val = ""
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = center_align
                if col_name in amount_cols and isinstance(val, (int, float)):
                    cell.number_format = accounting_format

            current_row += 1

        elif rtype == "total":
            total_row_nums.append(current_row)

            # Build SUM formulas using the tracked data row range
            for col_idx, col_name in enumerate(display_columns, 1):
                row_val = row_dict.get(col_name)

                # Use SUM formula for amount columns with more than 1 data row
                if (
                    col_name in amount_cols
                    and buyer_data_start is not None
                    and buyer_data_end is not None
                    and buyer_data_start <= buyer_data_end
                ):
                    col_letter = col_letter_map[col_name]
                    val = f"=SUM({col_letter}{buyer_data_start}:{col_letter}{buyer_data_end})"
                elif col_name == "Quantity":
                    val = "TOTAL"
                else:
                    val = "" if row_val is None or (isinstance(row_val, float) and pd.isna(row_val)) else row_val

                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = total_font
                cell.alignment = center_align

                # Borders only on the populated TOTAL-row columns
                if col_name in total_border_cols:
                    left = thick_side if col_name == "Quantity" else thin_side
                    right = thick_side if col_name == "Further Tax" else thin_side
                    cell.border = Border(
                        top=thick_side,
                        bottom=thick_side,
                        left=left,
                        right=right,
                    )

                if col_name in amount_cols:
                    cell.number_format = accounting_format

            # Merge Quantity + Rate cells for "TOTAL" label
            qty_col = display_columns.index("Quantity") + 1
            rate_col = display_columns.index("Rate") + 1
            ws.merge_cells(
                f"{get_column_letter(qty_col)}{current_row}:"
                f"{get_column_letter(rate_col)}{current_row}"
            )

            current_row += 1
            # Reset for the next buyer group
            buyer_data_start = None
            buyer_data_end = None

    # ── GRAND TOTAL row ─────────────────────────────────────────────────
    if total_row_nums:
        current_row += 1  # blank separator row

        for col_idx, col_name in enumerate(display_columns, 1):
            if col_name in amount_cols:
                col_letter = col_letter_map[col_name]
                # SUM of all buyer TOTAL rows
                cell_refs = ",".join(
                    f"{col_letter}{r}" for r in total_row_nums
                )
                val = f"=SUM({cell_refs})"
            elif col_name == "Quantity":
                val = "GRAND TOTAL"
            else:
                val = ""

            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = total_font
            cell.alignment = center_align

            # Same border style as regular TOTAL rows
            if col_name in total_border_cols:
                left = thick_side if col_name == "Quantity" else thin_side
                right = thick_side if col_name == "Further Tax" else thin_side
                cell.border = Border(
                    top=thick_side,
                    bottom=thick_side,
                    left=left,
                    right=right,
                )

            if col_name in amount_cols:
                cell.number_format = accounting_format

        # Merge Quantity + Rate cells for "GRAND TOTAL" label
        ws.merge_cells(
            f"{get_column_letter(qty_col)}{current_row}:"
            f"{get_column_letter(rate_col)}{current_row}"
        )

    # ── Column widths (looked up by column name, not hardcoded letter) ──
    column_widths = {
        "Invoice No.": 14,
        "Invoice Type": 16,
        "Invoice Date": 16,
        "Buyer Name": 32,
        "Seller Name": 18,
        "Quantity": 10,
        "Rate": 8,
        "Value of Sales Excluding Sales Tax": 38,
        "Sales Tax/ FED in ST Mode": 32,
        "Further Tax": 14,
    }
    for col_idx, col_name in enumerate(display_columns, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(col_name, 12)

    wb.save(output_path)


def convert(input_path: str, output_path: str, sort_by: str = "Buyer Name") -> None:
    """Orchestrator: read → clean → group → write.

    Parameters
    ----------
    input_path : str
        Path to the raw Excel file.
    output_path : str
        Path or buffer for the formatted output.
    sort_by : str
        Column to group/sort by — ``"Buyer Name"`` (default),
        ``"Seller Name"``, or ``"Tax Year"``.
    """
    df = read_input(input_path)
    df = clean_columns(df)
    rows = group_and_total(df, sort_by=sort_by)
    write_output(rows, output_path)
